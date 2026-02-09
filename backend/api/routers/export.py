"""
Export API endpoints.

Provides data export functionality for contracts and vendors in CSV and Excel formats.
"""
import csv
import io
import sqlite3
import logging
from typing import Optional
from datetime import datetime
from fastapi import APIRouter, Query, HTTPException, Request
from fastapi.responses import StreamingResponse

from ..dependencies import get_db
from ..config.constants import MAX_CONTRACT_VALUE

# Optional rate limiting - gracefully degrade if not available
try:
    from slowapi import Limiter
    from slowapi.util import get_remote_address
    limiter = Limiter(key_func=get_remote_address)
    RATE_LIMITING_ENABLED = True
except ImportError:
    limiter = None
    RATE_LIMITING_ENABLED = False


def rate_limit(limit_string: str):
    """
    Decorator factory for rate limiting.
    Returns a no-op decorator if slowapi is not installed.
    """
    if RATE_LIMITING_ENABLED and limiter:
        return limiter.limit(limit_string)
    else:
        def noop_decorator(func):
            return func
        return noop_decorator


logger = logging.getLogger(__name__)

# Export limits to prevent memory issues
MAX_EXPORT_ROWS = 100_000

router = APIRouter(prefix="/export", tags=["export"])


def sanitize_csv_cell(value):
    """Sanitize a CSV cell to prevent formula injection.

    CSV formula injection occurs when cells start with =, +, -, @, or tab/CR/LF
    followed by these characters. These can execute commands when opened in Excel.

    We prefix dangerous values with a single quote to neutralize them.
    """
    if value is None:
        return value

    # Convert to string if needed
    str_value = str(value) if not isinstance(value, str) else value

    # Check for formula injection patterns
    # Characters that can trigger formula execution in Excel/Google Sheets
    dangerous_prefixes = ('=', '+', '-', '@', '\t', '\r', '\n')

    if str_value and str_value[0] in dangerous_prefixes:
        # Prefix with single quote to neutralize (Excel displays it as text)
        return f"'{str_value}"

    return value


def generate_csv(rows, columns):
    """Generate CSV content from rows and columns with formula injection protection."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Sanitize column headers
    writer.writerow([sanitize_csv_cell(col) for col in columns])

    # Sanitize data rows
    for row in rows:
        writer.writerow([sanitize_csv_cell(cell) for cell in row])

    output.seek(0)
    return output


@router.get("/contracts/csv")
@rate_limit("10/minute")
def export_contracts_csv(
    request: Request,
    sector_id: Optional[int] = Query(None, ge=1, le=12, description="Filter by sector ID (1-12)"),
    year: Optional[int] = Query(None, ge=2002, le=2026, description="Filter by contract year"),
    institution_id: Optional[int] = Query(None, description="Filter by institution ID"),
    vendor_id: Optional[int] = Query(None, description="Filter by vendor ID"),
    risk_level: Optional[str] = Query(None, description="Filter by risk level (low/medium/high/critical)"),
    is_direct_award: Optional[bool] = Query(None, description="Filter direct awards"),
    is_single_bid: Optional[bool] = Query(None, description="Filter single-bid contracts"),
    min_amount: Optional[float] = Query(None, ge=0, description="Minimum contract amount"),
    max_amount: Optional[float] = Query(None, le=100_000_000_000, description="Maximum contract amount"),
    limit: int = Query(10000, ge=1, le=MAX_EXPORT_ROWS, description=f"Maximum rows to export (max {MAX_EXPORT_ROWS})"),
):
    """
    Export contracts as CSV file.

    Returns filtered contracts in CSV format for download.
    Maximum {MAX_EXPORT_ROWS} rows per export to prevent memory issues.
    """
    try:
        with get_db() as conn:
            cursor = conn.cursor()

            # Build WHERE clause
            conditions = ["COALESCE(c.amount_mxn, 0) <= ?"]
            params = [MAX_CONTRACT_VALUE]

            if sector_id is not None:
                conditions.append("c.sector_id = ?")
                params.append(sector_id)

            if year is not None:
                conditions.append("c.contract_year = ?")
                params.append(year)

            if institution_id is not None:
                conditions.append("c.institution_id = ?")
                params.append(institution_id)

            if vendor_id is not None:
                conditions.append("c.vendor_id = ?")
                params.append(vendor_id)

            if risk_level is not None:
                conditions.append("c.risk_level = ?")
                params.append(risk_level.lower())

            if is_direct_award is not None:
                conditions.append("c.is_direct_award = ?")
                params.append(1 if is_direct_award else 0)

            if is_single_bid is not None:
                conditions.append("c.is_single_bid = ?")
                params.append(1 if is_single_bid else 0)

            if min_amount is not None:
                conditions.append("c.amount_mxn >= ?")
                params.append(min_amount)

            if max_amount is not None:
                conditions.append("c.amount_mxn <= ?")
                params.append(max_amount)

            where_clause = " AND ".join(conditions)

            query = f"""
                SELECT
                    c.id,
                    c.contract_number,
                    c.procedure_number,
                    c.title,
                    c.description,
                    c.amount_mxn,
                    c.currency,
                    c.contract_date,
                    c.contract_year,
                    c.start_date,
                    c.end_date,
                    c.sector_id,
                    s.name_es as sector_name,
                    c.vendor_id,
                    v.name as vendor_name,
                    v.rfc as vendor_rfc,
                    c.institution_id,
                    i.name as institution_name,
                    i.institution_type,
                    c.procedure_type,
                    c.contract_type,
                    c.is_direct_award,
                    c.is_single_bid,
                    c.risk_score,
                    c.risk_level,
                    c.risk_factors
                FROM contracts c
                LEFT JOIN sectors s ON c.sector_id = s.id
                LEFT JOIN vendors v ON c.vendor_id = v.id
                LEFT JOIN institutions i ON c.institution_id = i.id
                WHERE {where_clause}
                ORDER BY c.contract_date DESC
                LIMIT ?
            """
            params.append(limit)
            cursor.execute(query, params)
            rows = cursor.fetchall()

            columns = [
                "id", "contract_number", "procedure_number", "title", "description",
                "amount_mxn", "currency", "contract_date", "contract_year",
                "start_date", "end_date", "sector_id", "sector_name",
                "vendor_id", "vendor_name", "vendor_rfc",
                "institution_id", "institution_name", "institution_type",
                "procedure_type", "contract_type",
                "is_direct_award", "is_single_bid",
                "risk_score", "risk_level", "risk_factors"
            ]

            # Convert rows to list format
            data_rows = [list(row) for row in rows]
            csv_output = generate_csv(data_rows, columns)

            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filters_str = ""
            if sector_id:
                filters_str += f"_sector{sector_id}"
            if year:
                filters_str += f"_year{year}"
            filename = f"contracts{filters_str}_{timestamp}.csv"

            return StreamingResponse(
                iter([csv_output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )

    except sqlite3.Error as e:
        logger.error(f"Database error in export_contracts_csv: {e}")
        raise HTTPException(status_code=500, detail="Database error occurred")


@router.get("/contracts/excel")
@rate_limit("10/minute")
def export_contracts_excel(
    request: Request,
    sector_id: Optional[int] = Query(None, ge=1, le=12, description="Filter by sector ID (1-12)"),
    year: Optional[int] = Query(None, ge=2002, le=2026, description="Filter by contract year"),
    institution_id: Optional[int] = Query(None, description="Filter by institution ID"),
    vendor_id: Optional[int] = Query(None, description="Filter by vendor ID"),
    risk_level: Optional[str] = Query(None, description="Filter by risk level (low/medium/high/critical)"),
    is_direct_award: Optional[bool] = Query(None, description="Filter direct awards"),
    is_single_bid: Optional[bool] = Query(None, description="Filter single-bid contracts"),
    min_amount: Optional[float] = Query(None, ge=0, description="Minimum contract amount"),
    max_amount: Optional[float] = Query(None, le=100_000_000_000, description="Maximum contract amount"),
    limit: int = Query(10000, ge=1, le=MAX_EXPORT_ROWS, description=f"Maximum rows to export (max {MAX_EXPORT_ROWS})"),
):
    """
    Export contracts as Excel file.

    Returns filtered contracts in Excel format for download.
    Maximum {MAX_EXPORT_ROWS} rows per export to prevent memory issues.
    """
    try:
        # Import openpyxl here to make it optional
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(
                status_code=501,
                detail="Excel export requires openpyxl. Install with: pip install openpyxl"
            )

        with get_db() as conn:
            cursor = conn.cursor()

            # Build WHERE clause (same as CSV)
            conditions = ["COALESCE(c.amount_mxn, 0) <= ?"]
            params = [MAX_CONTRACT_VALUE]

            if sector_id is not None:
                conditions.append("c.sector_id = ?")
                params.append(sector_id)

            if year is not None:
                conditions.append("c.contract_year = ?")
                params.append(year)

            if institution_id is not None:
                conditions.append("c.institution_id = ?")
                params.append(institution_id)

            if vendor_id is not None:
                conditions.append("c.vendor_id = ?")
                params.append(vendor_id)

            if risk_level is not None:
                conditions.append("c.risk_level = ?")
                params.append(risk_level.lower())

            if is_direct_award is not None:
                conditions.append("c.is_direct_award = ?")
                params.append(1 if is_direct_award else 0)

            if is_single_bid is not None:
                conditions.append("c.is_single_bid = ?")
                params.append(1 if is_single_bid else 0)

            if min_amount is not None:
                conditions.append("c.amount_mxn >= ?")
                params.append(min_amount)

            if max_amount is not None:
                conditions.append("c.amount_mxn <= ?")
                params.append(max_amount)

            where_clause = " AND ".join(conditions)

            query = f"""
                SELECT
                    c.id,
                    c.contract_number,
                    c.procedure_number,
                    c.title,
                    c.amount_mxn,
                    c.currency,
                    c.contract_date,
                    c.contract_year,
                    s.name_es as sector_name,
                    v.name as vendor_name,
                    v.rfc as vendor_rfc,
                    i.name as institution_name,
                    i.institution_type,
                    c.procedure_type,
                    c.is_direct_award,
                    c.is_single_bid,
                    c.risk_score,
                    c.risk_level
                FROM contracts c
                LEFT JOIN sectors s ON c.sector_id = s.id
                LEFT JOIN vendors v ON c.vendor_id = v.id
                LEFT JOIN institutions i ON c.institution_id = i.id
                WHERE {where_clause}
                ORDER BY c.contract_date DESC
                LIMIT ?
            """
            params.append(limit)
            cursor.execute(query, params)
            rows = cursor.fetchall()

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Contracts"

            # Headers
            headers = [
                "ID", "Contract Number", "Procedure Number", "Title",
                "Amount (MXN)", "Currency", "Contract Date", "Year",
                "Sector", "Vendor Name", "Vendor RFC",
                "Institution", "Institution Type", "Procedure Type",
                "Direct Award", "Single Bid", "Risk Score", "Risk Level"
            ]
            ws.append(headers)

            # Style header row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = cell.font.copy(bold=True)

            # Data rows - sanitize all cells to prevent formula injection
            for row in rows:
                sanitized_row = [sanitize_csv_cell(cell) for cell in row]
                ws.append(sanitized_row)

            # Auto-width columns (approximate)
            for i, col in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(i)].width = max(len(col) + 2, 12)

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filters_str = ""
            if sector_id:
                filters_str += f"_sector{sector_id}"
            if year:
                filters_str += f"_year{year}"
            filename = f"contracts{filters_str}_{timestamp}.xlsx"

            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )

    except sqlite3.Error as e:
        logger.error(f"Database error in export_contracts_excel: {e}")
        raise HTTPException(status_code=500, detail="Database error occurred")


@router.get("/vendors/csv")
@rate_limit("10/minute")
def export_vendors_csv(
    request: Request,
    sector_id: Optional[int] = Query(None, ge=1, le=12, description="Filter by primary sector"),
    min_contracts: Optional[int] = Query(None, ge=1, description="Minimum contract count"),
    min_value: Optional[float] = Query(None, ge=0, description="Minimum total contract value"),
    has_rfc: Optional[bool] = Query(None, description="Filter vendors with RFC"),
    limit: int = Query(10000, ge=1, le=MAX_EXPORT_ROWS, description=f"Maximum rows to export (max {MAX_EXPORT_ROWS})"),
):
    """
    Export vendors as CSV file.

    Returns vendors with aggregate statistics in CSV format for download.
    Maximum {MAX_EXPORT_ROWS} rows per export to prevent memory issues.
    """
    try:
        with get_db() as conn:
            cursor = conn.cursor()

            # Build WHERE clause
            conditions = ["1=1"]
            params = []

            if has_rfc is not None:
                if has_rfc:
                    conditions.append("v.rfc IS NOT NULL AND v.rfc != ''")
                else:
                    conditions.append("(v.rfc IS NULL OR v.rfc = '')")

            where_clause = " AND ".join(conditions)

            # HAVING conditions
            having_conditions = ["1=1"]
            having_params = []

            if min_contracts is not None:
                having_conditions.append("COUNT(c.id) >= ?")
                having_params.append(min_contracts)

            if min_value is not None:
                having_conditions.append("COALESCE(SUM(c.amount_mxn), 0) >= ?")
                having_params.append(min_value)

            if sector_id is not None:
                having_conditions.append("""
                    (SELECT sector_id FROM contracts
                     WHERE vendor_id = v.id
                     GROUP BY sector_id
                     ORDER BY COUNT(*) DESC LIMIT 1) = ?
                """)
                having_params.append(sector_id)

            having_clause = " AND ".join(having_conditions)

            query = f"""
                SELECT
                    v.id,
                    v.name,
                    v.rfc,
                    v.name_normalized,
                    COUNT(c.id) as total_contracts,
                    COALESCE(SUM(c.amount_mxn), 0) as total_value_mxn,
                    COALESCE(AVG(c.amount_mxn), 0) as avg_contract_value,
                    COALESCE(AVG(c.risk_score), 0) as avg_risk_score,
                    SUM(CASE WHEN c.risk_level IN ('high', 'critical') THEN 1 ELSE 0 END) as high_risk_count,
                    SUM(CASE WHEN c.is_direct_award = 1 THEN 1 ELSE 0 END) as direct_award_count,
                    SUM(CASE WHEN c.is_single_bid = 1 THEN 1 ELSE 0 END) as single_bid_count,
                    MIN(c.contract_year) as first_contract_year,
                    MAX(c.contract_year) as last_contract_year,
                    COUNT(DISTINCT c.institution_id) as total_institutions,
                    COUNT(DISTINCT c.sector_id) as sectors_count
                FROM vendors v
                LEFT JOIN contracts c ON v.id = c.vendor_id
                    AND COALESCE(c.amount_mxn, 0) <= ?
                WHERE {where_clause}
                GROUP BY v.id, v.name, v.rfc, v.name_normalized
                HAVING {having_clause}
                ORDER BY total_contracts DESC
                LIMIT ?
            """
            cursor.execute(query, [MAX_CONTRACT_VALUE] + params + having_params + [limit])
            rows = cursor.fetchall()

            columns = [
                "id", "name", "rfc", "name_normalized",
                "total_contracts", "total_value_mxn", "avg_contract_value",
                "avg_risk_score", "high_risk_count",
                "direct_award_count", "single_bid_count",
                "first_contract_year", "last_contract_year",
                "total_institutions", "sectors_count"
            ]

            # Convert rows to list format
            data_rows = [list(row) for row in rows]
            csv_output = generate_csv(data_rows, columns)

            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"vendors_{timestamp}.csv"

            return StreamingResponse(
                iter([csv_output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )

    except sqlite3.Error as e:
        logger.error(f"Database error in export_vendors_csv: {e}")
        raise HTTPException(status_code=500, detail="Database error occurred")
